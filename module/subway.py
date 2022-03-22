import openpyxl
import json

excel_file = '/workspaces/workspace_ai/03_BigdataPorject/subway.xlsx'
json_file = '/workspaces/workspace_ai/03_BigdataPorject/subway.json'

wb =openpyxl.load_workbook(excel_file, read_only=True)

sheet = wb.worksheets[0]

key_list = []
for col_num in range(1, sheet.max_column): # A ~ D
    key_list.append(sheet.cell(row=1, column=col_num).value)

# print(key_list)

data_dict = {}
key_index = 1
for row_num in range(2, sheet.max_row +1):
    tmp_dict = {}
    for col_num in range(1, sheet.max_column):
        val = sheet.cell(row=row_num, column=col_num).value
        tmp_dict[key_list[col_num - 1]] = val

    # print(tmp_dict)
    data_dict[tmp_dict[key_list[key_index]]] = tmp_dict

# print(data_dict)

wb.close()

with open(json_file, 'w', encoding='utf-8') as fp:
    json.dump(data_dict, fp, indent=4, ensure_ascii=False)